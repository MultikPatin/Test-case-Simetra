from collections.abc import Generator

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.utils.logger import create_logger
from src.exel_to_postgis.schemas.exel import ExelRow


class ExelExtractor:
    def __init__(self, filepath: str, buffer_size: int):
        self.__buffer_size = buffer_size
        self.__wb = load_workbook(filepath)
        self.__max_col = 6
        self.logger = create_logger("ExelExtractor")

    def __get_sheet(self, sheet_name: str = None) -> Worksheet:
        ws = None
        if sheet_name:
            try:
                ws = self.__wb["sheet_name"]
            except KeyError:
                self.logger.error(
                    "==> В файле не существует запрашиваемой вкладки"
                )
        else:
            ws = self.__wb.active
        return ws

    def extract(self) -> Generator:
        ws = self.__get_sheet()
        return self.__get_row_chunks_generator(ws)

    def __get_row_chunks_generator(self, ws: Worksheet) -> Generator:
        colum_headers = self.__get_colum_headers(ws)
        count = 2
        while count + self.__buffer_size <= ws.max_row:
            rows = self.__get_rows_values_by_position(
                ws, count, count + self.__buffer_size
            )
            self.logger.debug(
                f"В работе генератор по позициям {count}:{count + self.__buffer_size}"
            )
            count += self.__buffer_size
            rows = self.__get_row_in_frozen_set(rows, colum_headers)
            yield rows
        if ws.max_row - count > 0:
            rows = self.__get_rows_values_by_position(ws, count, ws.max_row)
            self.logger.debug(
                f"В работе генератор по позициям {count}:{ws.max_row}"
            )
            rows = self.__get_row_in_frozen_set(rows, colum_headers)
            yield rows

    def __get_rows_values_by_position(
        self, ws: Worksheet, min_row: int, max_row: int
    ) -> Generator:
        yield from ws.iter_rows(
            max_col=self.__max_col,
            min_row=min_row,
            max_row=max_row,
            values_only=True,
        )

    @staticmethod
    def __get_row_in_frozen_set(rows: Generator, headers: tuple) -> list:
        return [
            ExelRow(**{str(x): y for x, y in zip(headers, row)}) for row in rows
        ]

    def __get_colum_headers(self, ws: Worksheet) -> tuple:
        for row in ws.iter_rows(
            max_col=self.__max_col,
            min_row=1,
            max_row=1,
            values_only=True,
        ):
            return row
